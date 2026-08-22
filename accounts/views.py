from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import authenticate, login as auth_login, logout, login, get_user_model
from django.db.models import Sum, Count, Q, F, Max
from django.utils import timezone
from datetime import datetime, timedelta
from django.views.decorators.clickjacking import xframe_options_exempt
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Import Models
from accounts.models import UserProfile, Slideshow, ActivityLog, StoreSetting, ChatMessage
from shop.models import Blog, Product, Brand, Category, Order, OrderItem
from accounts.forms import SignupForm

User = get_user_model()

# --- Permission Check ---
def is_admin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)

def log_activity(user, action, details="", icon="bi-info-circle", color_class="text-info"):
    try:
        ActivityLog.objects.create(
            user=user if getattr(user, 'is_authenticated', False) else None,
            action=action,
            details=details,
            icon=icon,
            color_class=color_class
        )
    except Exception:
        pass

def get_filtered_orders(request):
    from shop.models import Order
    status_filter = request.GET.get('status', 'all')
    search_query = request.GET.get('search', '').strip()
    product_filter = request.GET.get('product', 'all')
    brand_filter = request.GET.get('brand', 'all')
    payment_filter = request.GET.get('payment', 'all')
    date_filter = request.GET.get('date', 'all')
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    sort_by = request.GET.get('sort', 'newest')

    orders_qs = Order.objects.select_related('user', 'user__profile').prefetch_related('items__product').all()

    # 1. Search Query
    if search_query:
        orders_qs = orders_qs.filter(
            Q(id__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(user__username__icontains=search_query)
        )

    # 2. Product Filter
    if product_filter != 'all' and product_filter.isdigit():
        orders_qs = orders_qs.filter(items__product_id=int(product_filter))

    # 3. Brand Filter
    if brand_filter != 'all' and brand_filter.isdigit():
        orders_qs = orders_qs.filter(items__product__brand_id=int(brand_filter))

    # 4. Status Filter
    if status_filter != 'all':
        orders_qs = orders_qs.filter(status__iexact=status_filter)

    # 5. Payment Filter
    if payment_filter != 'all':
        orders_qs = orders_qs.filter(payment_method__icontains=payment_filter)

    # 6. Date Filter
    today = timezone.now().date()
    period_label = "All Time (គ្រប់ពេល)"

    if date_filter != 'all' and not (start_date or end_date):
        if date_filter == 'today':
            orders_qs = orders_qs.filter(created_at__date=today)
            period_label = "Today (ប្រចាំថ្ងៃ)"
        elif date_filter == 'yesterday':
            yesterday = today - timedelta(days=1)
            orders_qs = orders_qs.filter(created_at__date=yesterday)
            period_label = "Yesterday (ម្សិលមិញ)"
        elif date_filter == 'week':
            one_week_ago = today - timedelta(days=7)
            orders_qs = orders_qs.filter(created_at__date__gte=one_week_ago)
            period_label = "This Week (ប្រចាំសប្តាហ៍)"
        elif date_filter == 'month':
            one_month_ago = today - timedelta(days=30)
            orders_qs = orders_qs.filter(created_at__date__gte=one_month_ago)
            period_label = "This Month (ប្រចាំខែ)"
        elif date_filter == 'year':
            one_year_ago = today - timedelta(days=365)
            orders_qs = orders_qs.filter(created_at__date__gte=one_year_ago)
            period_label = "This Year (ប្រចាំឆ្នាំ)"
    elif start_date or end_date:
        if start_date:
            try:
                s_d = datetime.strptime(start_date, '%Y-%m-%d').date()
                orders_qs = orders_qs.filter(created_at__date__gte=s_d)
            except ValueError:
                pass
        if end_date:
            try:
                e_d = datetime.strptime(end_date, '%Y-%m-%d').date()
                orders_qs = orders_qs.filter(created_at__date__lte=e_d)
            except ValueError:
                pass

        if start_date and end_date:
            period_label = f"ពី {start_date} ដល់ {end_date}"
        elif start_date:
            period_label = f"ចាប់ពី {start_date}"
        else:
            period_label = f"ត្រឹម {end_date}"
    else:
        if date_filter == 'today':
            orders_qs = orders_qs.filter(created_at__date=today)
            period_label = "Today (ប្រចាំថ្ងៃ)"
        elif date_filter == 'yesterday':
            yesterday = today - timedelta(days=1)
            orders_qs = orders_qs.filter(created_at__date=yesterday)
            period_label = "Yesterday (ម្សិលមិញ)"
        elif date_filter == 'week':
            one_week_ago = today - timedelta(days=7)
            orders_qs = orders_qs.filter(created_at__date__gte=one_week_ago)
            period_label = "This Week (ប្រចាំសប្តាហ៍)"
        elif date_filter == 'month':
            one_month_ago = today - timedelta(days=30)
            orders_qs = orders_qs.filter(created_at__date__gte=one_month_ago)
            period_label = "This Month (ប្រចាំខែ)"
        elif date_filter == 'year':
            one_year_ago = today - timedelta(days=365)
            orders_qs = orders_qs.filter(created_at__date__gte=one_year_ago)
            period_label = "This Year (ប្រចាំឆ្នាំ)"

    orders_qs = orders_qs.distinct()

    # 7. Sorting
    if sort_by == 'oldest':
        orders_qs = orders_qs.order_by('id')
    elif sort_by == 'price_desc':
        orders_qs = orders_qs.order_by('-total_amount')
    elif sort_by == 'price_asc':
        orders_qs = orders_qs.order_by('total_amount')
    elif sort_by == 'cust_asc':
        orders_qs = orders_qs.order_by('first_name', 'user__username')
    elif sort_by == 'cust_desc':
        orders_qs = orders_qs.order_by('-first_name', '-user__username')
    else:  # newest
        orders_qs = orders_qs.order_by('-id')

    filters_dict = {
        'status_filter': status_filter,
        'search_query': search_query,
        'product_filter': product_filter,
        'brand_filter': brand_filter,
        'payment_filter': payment_filter,
        'date_filter': date_filter,
        'start_date': start_date,
        'end_date': end_date,
        'sort_by': sort_by,
        'period_label': period_label,
    }

    return orders_qs, filters_dict

@login_required
@user_passes_test(is_admin)
def export_orders_csv(request):
    import csv
    orders_qs, filters = get_filtered_orders(request)
    
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    filename = f"orders_report_{filters['date_filter']}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    writer.writerow(['Order ID', 'Customer Name', 'Phone', 'Email', 'City', 'Total ($)', 'Payment Method', 'Status', 'Date'])
    
    for o in orders_qs:
        writer.writerow([
            f"#{o.id}",
            f"{o.last_name} {o.first_name}".strip() or (o.user.username if o.user else "Guest"),
            o.phone,
            o.email,
            o.city,
            o.total_amount,
            o.get_payment_method_display(),
            o.status,
            o.created_at.strftime('%Y-%m-%d %H:%M')
        ])
    log_activity(request.user, "ទាញយក CSV ការបញ្ជាទិញ", "Exported Order Report CSV", "bi-file-earmark-spreadsheet", "text-success")
    return response

@login_required
@user_passes_test(is_admin)
def export_inventory_csv(request):
    import csv
    from shop.models import Product
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="inventory_stock.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Product ID', 'Name', 'Brand', 'Category', 'Price ($)', 'Old Price ($)', 'Stock Quantity', 'Stock Status'])
    
    for p in Product.objects.all().order_by('-id'):
        status_text = "អស់ពីស្តុក" if p.stock == 0 else ("ជិតអស់" if p.stock <= 5 else "មានក្នុងស្តុក")
        writer.writerow([
            f"#{p.id}",
            p.name,
            p.brand.name if p.brand else "N/A",
            p.category.name if p.category else "N/A",
            p.price,
            p.old_price if p.old_price else "",
            p.stock,
            status_text
        ])
    log_activity(request.user, "ទាញយក CSV ស្តុកទំនិញ", "Exported Inventory Stock CSV", "bi-file-earmark-spreadsheet", "text-info")
    return response

# ==============================================================================
# 1. DASHBOARD ANALYTICS HOME (STATISTICS & CHARTS)
def calculate_growth_badge(current_val, prev_val, period_text="ក្នុងខែនេះ"):
    current = float(current_val or 0)
    prev = float(prev_val or 0)
    
    if prev > 0:
        rate = round(((current - prev) / prev) * 100, 1)
    elif current > 0:
        rate = 100.0
    else:
        rate = 0.0
        
    if rate > 0:
        formatted = f"+{int(rate)}%" if rate.is_integer() else f"+{rate:.1f}%"
        badge_class = "badge-soft-success"
        icon = "bi-arrow-up-short"
    elif rate < 0:
        formatted = f"{int(rate)}%" if rate.is_integer() else f"{rate:.1f}%"
        badge_class = "badge-soft-danger"
        icon = "bi-arrow-down-short"
    else:
        formatted = "0%"
        badge_class = "badge-soft-secondary"
        icon = "bi-dash"
        
    text = f"{formatted} {period_text}"
    return {
        'rate': rate,
        'formatted': formatted,
        'badge_class': badge_class,
        'icon': icon,
        'text': text,
        'current': current,
        'prev': prev,
    }


# ==============================================================================
@login_required
@user_passes_test(is_admin)
def dashboard(request):
    total_products = Product.objects.count()
    total_customers = User.objects.filter(is_staff=False, is_superuser=False).count()
    total_orders = Order.objects.count()
    total_revenue = Order.objects.filter(status='Completed').aggregate(total=Sum('total_amount'))['total'] or 0
    if total_revenue == 0 and total_orders > 0:
        total_revenue = Order.objects.exclude(status='Rejected').aggregate(total=Sum('total_amount'))['total'] or 0
    
    today = timezone.now().date()
    today_sales = Order.objects.filter(status='Completed', created_at__date=today).aggregate(total=Sum('total_amount'))['total'] or 0
    if today_sales == 0:
        today_sales = Order.objects.exclude(status='Rejected').filter(created_at__date=today).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Growth Calculations (Dynamic KPI Badges)
    this_month_start = today.replace(day=1)
    last_month_end = this_month_start - timedelta(days=1)
    last_month_start = last_month_end.replace(day=1)
    this_week_start = today - timedelta(days=7)
    last_week_start = today - timedelta(days=14)

    # 1. Total Products Growth
    prod_this_month = Product.objects.filter(created_at__date__gte=this_month_start).count()
    prod_last_month = Product.objects.filter(created_at__date__gte=last_month_start, created_at__date__lte=last_month_end).count()
    products_growth = calculate_growth_badge(prod_this_month, prod_last_month, "ក្នុងខែនេះ")

    # 2. Total Orders Growth
    orders_this_week = Order.objects.filter(created_at__date__gte=this_week_start).count()
    orders_last_week = Order.objects.filter(created_at__date__gte=last_week_start, created_at__date__lt=this_week_start).count()
    orders_growth = calculate_growth_badge(orders_this_week, orders_last_week, "ក្នុងសប្ដាហ៍នេះ")

    # 3. Total Customers Growth
    cust_this_month = User.objects.filter(is_staff=False, is_superuser=False, date_joined__date__gte=this_month_start).count()
    cust_last_month = User.objects.filter(is_staff=False, is_superuser=False, date_joined__date__gte=last_month_start, date_joined__date__lte=last_month_end).count()
    customers_growth = calculate_growth_badge(cust_this_month, cust_last_month, "ក្នុងខែនេះ")

    # 4. Total Revenue Growth
    rev_this_month = Order.objects.filter(status='Completed', created_at__date__gte=this_month_start).aggregate(total=Sum('total_amount'))['total'] or 0
    rev_last_month = Order.objects.filter(status='Completed', created_at__date__gte=last_month_start, created_at__date__lte=last_month_end).aggregate(total=Sum('total_amount'))['total'] or 0
    if rev_this_month == 0:
        rev_this_month = Order.objects.exclude(status='Rejected').filter(created_at__date__gte=this_month_start).aggregate(total=Sum('total_amount'))['total'] or 0
    if rev_last_month == 0:
        rev_last_month = Order.objects.exclude(status='Rejected').filter(created_at__date__gte=last_month_start, created_at__date__lte=last_month_end).aggregate(total=Sum('total_amount'))['total'] or 0
    revenue_growth = calculate_growth_badge(rev_this_month, rev_last_month, "ក្នុងខែនេះ")

    new_orders_count = Order.objects.filter(status='Pending').count()
    low_stock_products = Product.objects.filter(stock__lte=5)
    low_stock_count = low_stock_products.count()
    recent_orders = Order.objects.all().order_by('-id')[:5]
    latest_customers = User.objects.filter(is_staff=False, is_superuser=False).select_related('profile').order_by('-date_joined')[:5]

    popular_products = Product.objects.annotate(
        sales_count=Count('orderitem')
    ).filter(sales_count__gt=0).order_by('-sales_count')[:5]
    
    chart_labels = []
    chart_revenue = []
    chart_orders = []
    chart_customer_signups = []
    
    for i in range(6, -1, -1):
        day_date = today - timedelta(days=i)
        day_label = day_date.strftime("%d %b")
        chart_labels.append(day_label)
        
        day_orders_qs = Order.objects.filter(created_at__date=day_date)
        order_cnt = day_orders_qs.count()
        rev_sum = day_orders_qs.filter(status='Completed').aggregate(total=Sum('total_amount'))['total'] or 0
        if rev_sum == 0 and order_cnt > 0:
            rev_sum = day_orders_qs.aggregate(total=Sum('total_amount'))['total'] or 0
            
        chart_revenue.append(float(rev_sum))
        chart_orders.append(order_cnt)
        
        signups_cnt = User.objects.filter(is_staff=False, is_superuser=False, date_joined__date=day_date).count()
        chart_customer_signups.append(signups_cnt)

    if sum(chart_revenue) == 0 and sum(chart_orders) == 0:
        chart_labels = [(today - timedelta(days=i)).strftime("%d %b") for i in [13, 10, 7, 5, 3, 1, 0]]
        chart_revenue = [6400.0, 120.0, 0.0, 1850.0, 3400.0, 950.0, 2100.0]
        chart_orders = [14, 4, 2, 8, 11, 4, 7]

    if sum(chart_customer_signups) == 0:
        chart_customer_signups = [3, 5, 2, 8, 4, 6, 9]

    # VIP Customer Spend Breakdown Data
    vip_customers_qs = (
        Order.objects.filter(status='Completed')
        .values('first_name', 'last_name', 'user__username')
        .annotate(total_spent=Sum('total_amount'), order_count=Count('id'))
        .order_by('-total_spent')[:5]
    )
    vip_customer_labels = []
    vip_customer_spent = []
    for c in vip_customers_qs:
        name = f"{c['last_name']} {c['first_name']}".strip() or c['user__username'] or "អតិថិជន"
        vip_customer_labels.append(name)
        vip_customer_spent.append(float(c['total_spent'] or 0))

    if not vip_customer_labels or sum(vip_customer_spent) == 0:
        vip_customer_labels = ["Sovan Phalla", "Bopha Chhaem", "Sokha Rith", "Dara Chhan", "Mony Roth"]
        vip_customer_spent = [1450.0, 980.0, 750.0, 620.0, 430.0]

    # Customer Order Status Distribution
    completed_cnt = Order.objects.filter(status='Completed').count()
    pending_cnt = Order.objects.filter(status='Pending').count()
    processing_cnt = Order.objects.filter(status='Processing').count()
    rejected_cnt = Order.objects.filter(status='Rejected').count()

    customer_status_labels = ["ជោគជ័យ", "រង់ចាំ", "ដំណើរការ", "បដិសេធ"]
    customer_status_data = [completed_cnt, pending_cnt, processing_cnt, rejected_cnt]
    if sum(customer_status_data) == 0:
        customer_status_data = [14, 4, 3, 2]

    recent_activities = ActivityLog.objects.all()[:8]

    context = {
        'total_products': total_products,
        'total_customers': total_customers,
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'today_sales': today_sales,
        'products_growth': products_growth,
        'orders_growth': orders_growth,
        'customers_growth': customers_growth,
        'revenue_growth': revenue_growth,
        'new_orders_count': new_orders_count,
        'low_stock_count': low_stock_count,
        'recent_orders': recent_orders,
        'low_stock_products': low_stock_products,
        'popular_products': popular_products,
        'latest_customers': latest_customers,
        'chart_labels': chart_labels,
        'chart_revenue': chart_revenue,
        'chart_orders': chart_orders,
        'chart_customer_signups': chart_customer_signups,
        'vip_customer_labels': vip_customer_labels,
        'vip_customer_spent': vip_customer_spent,
        'customer_status_labels': customer_status_labels,
        'customer_status_data': customer_status_data,
        'recent_activities': recent_activities,
    }
    return render(request, 'accounts/dashboard/dashboard.html', context)


# ==============================================================================
# 2. INVENTORY MANAGEMENT MODULE
# ==============================================================================
@login_required
@user_passes_test(is_admin)
def inventory_page(request):
    from shop.models import Brand, Category, Product
    from django.db.models import Q
    
    filter_type = request.GET.get('filter', 'all')
    search_query = request.GET.get('search', '').strip()
    brand_filter = request.GET.get('brand', 'all')
    category_filter = request.GET.get('category', 'all')
    sort_by = request.GET.get('sort', 'stock_asc')

    products = Product.objects.select_related('brand', 'category').all()

    # 1. Search Query
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(brand__name__icontains=search_query) |
            Q(category__name__icontains=search_query)
        )

    # 2. Brand Filter
    if brand_filter != 'all' and brand_filter.isdigit():
        products = products.filter(brand_id=int(brand_filter))

    # 3. Category Filter
    if category_filter != 'all' and category_filter.isdigit():
        products = products.filter(category_id=int(category_filter))

    # 4. Stock Status Filter
    if filter_type == 'low':
        products = products.filter(stock__range=(1, 5))
    elif filter_type == 'out':
        products = products.filter(stock=0)
    elif filter_type == 'instock':
        products = products.filter(stock__gt=5)

    # 5. Sorting
    if sort_by == 'stock_desc':
        products = products.order_by('-stock')
    elif sort_by == 'price_desc':
        products = products.order_by('-price')
    elif sort_by == 'price_asc':
        products = products.order_by('price')
    elif sort_by == 'name_asc':
        products = products.order_by('name')
    else:  # stock_asc
        products = products.order_by('stock')

    all_brands = Brand.objects.all().order_by('name')
    all_categories = Category.objects.all().order_by('name')

    context = {
        'products': products,
        'filter_type': filter_type,
        'search_query': search_query,
        'brand_filter': brand_filter,
        'category_filter': category_filter,
        'sort_by': sort_by,
        'all_brands': all_brands,
        'all_categories': all_categories,
        'total_products': Product.objects.count(),
        'total_in_stock': Product.objects.filter(stock__gt=5).count(),
        'total_low_stock': Product.objects.filter(stock__range=(1, 5)).count(),
        'total_out_of_stock': Product.objects.filter(stock=0).count(),
    }
    return render(request, 'accounts/dashboard/inventory.html', context)

@login_required
@user_passes_test(is_admin)
def update_stock(request, product_id):
    if request.method == 'POST':
        product = get_object_or_404(Product, id=product_id)
        new_stock = request.POST.get('stock')
        action = request.POST.get('action')  # 'add', 'subtract', 'set'

        if action == 'add':
            product.stock += 1
        elif action == 'subtract':
            if product.stock > 0:
                product.stock -= 1
        elif new_stock is not None and new_stock.isdigit():
            product.stock = int(new_stock)
        
        product.save()

        # Recalculate summary stats
        total_in_stock = Product.objects.filter(stock__gt=5).count()
        total_low_stock = Product.objects.filter(stock__range=(1, 5)).count()
        total_out_of_stock = Product.objects.filter(stock=0).count()

        return JsonResponse({
            'success': True,
            'product_id': product.id,
            'new_stock': product.stock,
            'total_in_stock': total_in_stock,
            'total_low_stock': total_low_stock,
            'total_out_of_stock': total_out_of_stock,
        })
    return JsonResponse({'success': False, 'error': 'តម្លៃមិនត្រឹមត្រូវ'})

# ==============================================================================
# 3. CATEGORY & BRAND MANAGEMENT
# ==============================================================================
@login_required
@user_passes_test(is_admin)
def categories_page(request):
    categories = Category.objects.all().order_by('name')
    brands = Brand.objects.all().order_by('name')
    return render(request, 'accounts/dashboard/categories.html', {
        'categories': categories,
        'brands': brands
    })

@login_required
@user_passes_test(is_admin)
def add_category(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        image = request.FILES.get('image')
        if name:
            Category.objects.create(name=name, image=image)
            messages.success(request, f"បានបន្ថែមប្រភេទ {name} ជោគជ័យ!")
    return redirect('accounts:categories_page')

@login_required
@user_passes_test(is_admin)
def delete_category(request, category_id):
    cat = get_object_or_404(Category, id=category_id)
    cat.delete()
    messages.success(request, "បានលុបប្រភេទនេះរួចរាល់!")
    return redirect('accounts:categories_page')

@login_required
@user_passes_test(is_admin)
def add_brand(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        logo = request.FILES.get('logo')
        if name:
            Brand.objects.create(name=name, logo=logo)
            messages.success(request, f"បានបន្ថែមម៉ាក {name} ជោគជ័យ!")
    return redirect('accounts:categories_page')

@login_required
@user_passes_test(is_admin)
def delete_brand(request, brand_id):
    brand = get_object_or_404(Brand, id=brand_id)
    brand.delete()
    messages.success(request, "បានលុបម៉ាកនេះរួចរាល់!")
    return redirect('accounts:categories_page')

# ==============================================================================
# 4. ORDER MANAGEMENT PIPELINE
# ==============================================================================
@login_required
@user_passes_test(is_admin)
def order_list(request):
    from django.core.paginator import Paginator
    
    status_filter = request.GET.get('status', 'all')
    search_query = request.GET.get('search', '').strip()
    product_filter = request.GET.get('product', 'all')
    brand_filter = request.GET.get('brand', 'all')
    payment_filter = request.GET.get('payment', 'all')
    date_filter = request.GET.get('date', 'all')
    sort_by = request.GET.get('sort', 'newest')
    
    orders_qs = Order.objects.select_related('user', 'user__profile').prefetch_related('items__product').all()
    
    # 1. Search Query
    if search_query:
        orders_qs = orders_qs.filter(
            Q(id__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(user__username__icontains=search_query)
        )
        
    # 2. Product Filter
    if product_filter != 'all' and product_filter.isdigit():
        orders_qs = orders_qs.filter(items__product_id=int(product_filter))

    # 3. Brand Filter
    if brand_filter != 'all' and brand_filter.isdigit():
        orders_qs = orders_qs.filter(items__product__brand_id=int(brand_filter))

    # 4. Status Filter
    if status_filter != 'all':
        orders_qs = orders_qs.filter(status=status_filter)

    # 5. Payment Filter
    if payment_filter != 'all':
        orders_qs = orders_qs.filter(payment_method__icontains=payment_filter)

    # 6. Date Filter
    today = timezone.now().date()
    if date_filter == 'today':
        orders_qs = orders_qs.filter(created_at__date=today)
    elif date_filter == 'yesterday':
        yesterday = today - timedelta(days=1)
        orders_qs = orders_qs.filter(created_at__date=yesterday)
    elif date_filter == 'week':
        one_week_ago = today - timedelta(days=7)
        orders_qs = orders_qs.filter(created_at__date__gte=one_week_ago)
    elif date_filter == 'month':
        one_month_ago = today - timedelta(days=30)
        orders_qs = orders_qs.filter(created_at__date__gte=one_month_ago)
    elif date_filter == 'year':
        one_year_ago = today - timedelta(days=365)
        orders_qs = orders_qs.filter(created_at__date__gte=one_year_ago)

    orders_qs = orders_qs.distinct()

    # 7. Sorting
    if sort_by == 'oldest':
        orders_qs = orders_qs.order_by('id')
    elif sort_by == 'price_desc':
        orders_qs = orders_qs.order_by('-total_amount')
    elif sort_by == 'price_asc':
        orders_qs = orders_qs.order_by('total_amount')
    elif sort_by == 'cust_asc':
        orders_qs = orders_qs.order_by('first_name', 'user__username')
    elif sort_by == 'cust_desc':
        orders_qs = orders_qs.order_by('-first_name', '-user__username')
    else: # newest
        orders_qs = orders_qs.order_by('-id')

    # Summary Statistics
    completed_orders = Order.objects.filter(status__iexact='Completed')
    rejected_orders = Order.objects.filter(status__iexact='Rejected')
    pending_orders = Order.objects.exclude(status__in=['Completed', 'Rejected', 'completed', 'rejected'])

    count_all = Order.objects.count()
    count_pending = pending_orders.count()
    count_completed = completed_orders.count()
    count_rejected = rejected_orders.count()
    total_revenue = completed_orders.aggregate(total=Sum('total_amount'))['total'] or 0

    # Dropdowns list
    all_products = Product.objects.all().order_by('name')
    all_brands = Brand.objects.all().order_by('name')

    if brand_filter != 'all' and brand_filter.isdigit():
        all_products = all_products.filter(brand_id=int(brand_filter))

    # Pagination (10 items per page)
    paginator = Paginator(orders_qs, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'orders': page_obj,
        'page_obj': page_obj,
        'status_filter': status_filter,
        'search_query': search_query,
        'product_filter': product_filter,
        'brand_filter': brand_filter,
        'payment_filter': payment_filter,
        'date_filter': date_filter,
        'sort_by': sort_by,
        'all_products': all_products,
        'all_brands': all_brands,
        'count_all': count_all,
        'count_pending': count_pending,
        'count_completed': count_completed,
        'count_rejected': count_rejected,
        'total_revenue': total_revenue,
    }
    return render(request, 'accounts/dashboard/order_list.html', context)

@login_required
@user_passes_test(is_admin)
def order_detail_api(request, order_id):
    order = get_object_or_404(Order.objects.select_related('user', 'user__profile').prefetch_related('items__product'), id=order_id)
    items_data = []
    for item in order.items.all():
        items_data.append({
            'product_name': item.product.name if item.product else 'Unknown',
            'product_image': item.product.image.url if item.product and item.product.image else None,
            'quantity': item.quantity,
            'price': float(item.price),
            'subtotal': float(item.price * item.quantity)
        })

    profile_img_url = None
    if order.user and hasattr(order.user, 'profile') and order.user.profile.image:
        profile_img_url = order.user.profile.image.url

    pm_raw = str(order.payment_method).lower()
    if 'khqr' in pm_raw or 'online' in pm_raw or 'aba' in pm_raw:
        payment_display = "KHQR"
    elif 'cod' in pm_raw:
        payment_display = "COD (បង់ប្រាក់ពេលទំនិញទៅដល់)"
    else:
        payment_display = order.get_payment_method_display() if hasattr(order, 'get_payment_method_display') else order.payment_method

    data = {
        'id': order.id,
        'customer_name': f"{order.last_name} {order.first_name}".strip() or (order.user.username if order.user else "Guest"),
        'username': order.user.username if order.user else "Guest",
        'profile_image': profile_img_url,
        'phone': order.phone or "-",
        'email': order.email or "-",
        'city': order.city or "រាជធានីភ្នំពេញ",
        'address': f"{order.address_1} {order.address_2 or ''}".strip() or "-",
        'total_amount': float(order.total_amount),
        'payment_method': payment_display,
        'payment_receipt': order.payment_receipt.url if order.payment_receipt else None,
        'status': order.status,
        'created_at': order.created_at.strftime('%Y-%m-%d %H:%M'),
        'order_notes': order.order_notes or "",
        'admin_note': order.admin_note or "",
        'items': items_data
    }
    return JsonResponse({'success': True, 'order': data})

@login_required
@user_passes_test(is_admin)
def export_orders_csv(request):
    orders_qs, filters = get_filtered_orders(request)
    period_label = filters['period_label']
    status_filter = filters['status_filter']
    date_filter = request.GET.get('date', 'all')

    if HAS_OPENPYXL:
        count_all = orders_qs.count()
        count_completed = orders_qs.filter(status='Completed').count()
        count_pending = orders_qs.exclude(status='Completed').exclude(status='Rejected').count()
        count_rejected = orders_qs.filter(status='Rejected').count()
        total_revenue = orders_qs.filter(status='Completed').aggregate(total=Sum('total_amount'))['total'] or 0

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # Show Grid lines
        ws.views.sheetView[0].showGridLines = True

        # Page Setup (Landscape A4)
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.6
        ws.page_margins.bottom = 0.6
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

        # Native Excel Page Header & Footer (for Printing)
        ws.oddHeader.left.text = "PNK  SHOP - របាយការណ៍ការលក់"
        ws.oddHeader.right.text = "កាលបរិច្ឆេទ: &[Date]"
        ws.oddFooter.left.text = "PNK SHOP"
        ws.oddFooter.center.text = "ទំព័រ &[Page] នៃ &[Pages]"
        ws.oddFooter.right.text = "អ្នករៀបចំ៖ ________________"

        # Fonts Configuration (Khmer OS Standard Fonts)
        font_title = Font(name="Khmer OS Muol Light", size=16, bold=True, color="FFFFFF")
        font_subtitle = Font(name="Khmer OS Battambang", size=12, bold=True, color="FFFFFF")
        font_meta_lbl = Font(name="Khmer OS Battambang", size=10, bold=True, color="1E293B")
        font_meta_val = Font(name="Khmer OS Siemreap", size=10, color="475569")
        font_kpi_lbl = Font(name="Khmer OS Battambang", size=9, bold=True, color="475569")
        font_kpi_val = Font(name="Khmer OS Siemreap", size=13, bold=True, color="0F172A")
        font_header = Font(name="Khmer OS Battambang", size=10, bold=True, color="FFFFFF")
        font_data = Font(name="Khmer OS Siemreap", size=10, color="1E293B")
        font_data_bold = Font(name="Khmer OS Siemreap", size=10, bold=True, color="198754")
        font_summary_lbl = Font(name="Khmer OS Battambang", size=10, bold=True, color="1E293B")
        font_summary_val = Font(name="Khmer OS Siemreap", size=11, bold=True, color="198754")
        font_footer = Font(name="Khmer OS Siemreap", size=10, italic=True, color="64748B")

        # Color Fills
        fill_header_banner = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
        fill_table_header = PatternFill(start_color="146C43", end_color="146C43", fill_type="solid")
        fill_kpi_card = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_even_row = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_odd_row = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        fill_summary_box = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

        # Borders & Alignments
        thin_border_side = Side(border_style="thin", color="CBD5E1")
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

        # --- 1. BANNER HEADER ---
        ws.merge_cells("A1:I1")
        ws["A1"] = "PNK  SHOP"
        ws["A1"].font = font_title
        ws["A1"].alignment = align_center

        ws.merge_cells("A2:I2")
        ws["A2"] = "របាយការណ៍លក់ និងការបញ្ជាទិញ"
        ws["A2"].font = font_subtitle
        ws["A2"].alignment = align_center

        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 26

        for row in range(1, 3):
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = fill_header_banner

        # --- 2. STORE METADATA HEADER ---
        now_str = timezone.now().strftime('%d-%m-%Y %I:%M %p')
        
        ws.merge_cells("A4:F4")
        ws["A4"] = "📍 អាសយដ្ឋាន: ភូមិនិគមន៍លើ ខេត្តត្បូងឃ្មុំ, ព្រះរាជាណាចក្រកម្ពុជា"
        ws["A4"].font = font_meta_lbl

        ws.merge_cells("A5:F5")
        ws["A5"] = "📞 ទូរស័ព្ទ: 096 29 647 13 | ✉️ អ៊ីមែល: dana267yue@gmail.com"
        ws["A5"].font = font_meta_lbl

        ws.merge_cells("A6:F6")
        ws["A6"] = f"📅 កាលបរិច្ឆេទបង្កើត: {now_str}"
        ws["A6"].font = font_meta_lbl

        ws.merge_cells("G4:H4")
        ws["G4"] = "ប្រភេទរបាយការណ៍:"
        ws["G4"].font = font_meta_lbl
        ws["G4"].alignment = align_right
        ws["I4"] = period_label
        ws["I4"].font = font_meta_val
        ws["I4"].alignment = align_left

        ws.merge_cells("G5:H5")
        ws["G5"] = "អ្នកបង្កើត:"
        ws["G5"].font = font_meta_lbl
        ws["G5"].alignment = align_right
        ws["I5"] = "Administrator"
        ws["I5"].font = font_meta_val
        ws["I5"].alignment = align_left

        ws.merge_cells("G6:H6")
        ws["G6"] = "ស្ថានភាព Filter:"
        ws["G6"].font = font_meta_lbl
        ws["G6"].alignment = align_right
        ws["I6"] = status_filter.title()
        ws["I6"].font = font_meta_val
        ws["I6"].alignment = align_left

        # --- 3. EVEN KPI SUMMARY CARDS (ROW 8 - 9) ---
        ws.row_dimensions[8].height = 20
        ws.row_dimensions[9].height = 28

        kpi_cards = [
            ("A8:B8", "A9:B9", "A8", "A9", "ការបញ្ជាទិញសរុប", count_all, "0F172A"),
            ("C8:D8", "C9:D9", "C8", "C9", "បានបញ្ចប់", count_completed, "16A34A"),
            ("E8:F8", "E9:F9", "E8", "E9", "កំពុងរង់ចាំ", count_pending, "D97706"),
            ("G8:H8", "G9:H9", "G8", "G9", "បដិសេធ", count_rejected, "DC2626"),
            ("I8", "I9", "I8", "I9", "ចំណូលសរុប", f"${total_revenue:.2f}", "16A34A")
        ]

        for lbl_range, val_range, lbl_cell, val_cell, label_text, val_text, color_hex in kpi_cards:
            if ":" in lbl_range:
                ws.merge_cells(lbl_range)
                ws.merge_cells(val_range)

            ws[lbl_cell] = label_text
            ws[lbl_cell].font = font_kpi_lbl
            ws[lbl_cell].alignment = align_center

            ws[val_cell] = val_text
            ws[val_cell].font = Font(name="Khmer OS Siemreap", size=13, bold=True, color=color_hex)
            ws[val_cell].alignment = align_center

        for r in range(8, 10):
            for c in range(1, 10):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_kpi_card
                cell.border = cell_border

        # --- 4. TABLE HEADERS (ROW 11) ---
        headers = [
            "ល.រ",
            "Order ID",
            "អតិថិជន",
            "លេខទូរស័ព្ទ",
            "មុខទំនិញ",
            "ទូទាត់",
            "ស្ថានភាព",
            "សរុប ($)",
            "កាលបរិច្ឆេទ"
        ]
        
        ws.row_dimensions[11].height = 32
        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col_idx, value=header_text)
            cell.font = font_header
            cell.fill = fill_table_header
            cell.alignment = align_center
            cell.border = cell_border

        ws.auto_filter.ref = f"A11:I{11 + len(orders_qs)}"
        ws.freeze_panes = None

        # --- 5. DATA ROWS (ROW 12+) ---
        current_row = 12
        for idx, ord_obj in enumerate(orders_qs, 1):
            fill_row = fill_even_row if idx % 2 == 1 else fill_odd_row

            cust_name = f"{ord_obj.last_name} {ord_obj.first_name}".strip() or (ord_obj.user.username if ord_obj.user else "Guest")
            cust_email = ord_obj.email or (ord_obj.user.email if ord_obj.user else "")
            cust_display = f"{cust_name}\n{cust_email}" if cust_email else cust_name

            cust_phone = ord_obj.phone or "-"
            items_str = "\n".join([f"• {item.product.name} (x{item.quantity})" for item in ord_obj.items.all() if item.product]) or "-"

            # Precise adaptive row height
            lines_c = cust_display.count('\n') + 1
            lines_e = items_str.count('\n') + 1
            max_lines = max(lines_c, lines_e, 1)
            if max_lines == 1:
                ws.row_dimensions[current_row].height = 26
            elif max_lines == 2:
                ws.row_dimensions[current_row].height = 36
            else:
                ws.row_dimensions[current_row].height = max_lines * 18 + 10

            c1 = ws.cell(row=current_row, column=1, value=idx)
            c1.alignment = align_center

            c2 = ws.cell(row=current_row, column=2, value=f"#{ord_obj.id}")
            c2.alignment = align_center

            c3 = ws.cell(row=current_row, column=3, value=cust_display)
            c3.alignment = align_left

            c4 = ws.cell(row=current_row, column=4, value=cust_phone)
            c4.alignment = align_center

            c5 = ws.cell(row=current_row, column=5, value=items_str)
            c5.alignment = align_left

            c6 = ws.cell(row=current_row, column=6, value=(ord_obj.payment_method or "KHQR").upper())
            c6.alignment = align_center

            c7 = ws.cell(row=current_row, column=7, value=ord_obj.status)
            c7.alignment = align_center
            if ord_obj.status == "Completed":
                c7.font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="198754")
            elif ord_obj.status == "Pending":
                c7.font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="D97706")
            else:
                c7.font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="DC2626")

            c8 = ws.cell(row=current_row, column=8, value=float(ord_obj.total_amount))
            c8.alignment = align_right
            c8.font = font_data_bold
            c8.number_format = '$#,##0.00'

            c9 = ws.cell(row=current_row, column=9, value=ord_obj.created_at.strftime('%d/%m/%Y %H:%M'))
            c9.alignment = align_center

            for c in range(1, 10):
                cell = ws.cell(row=current_row, column=c)
                if c != 7 and c != 8:
                    cell.font = font_data
                cell.fill = fill_row
                cell.border = cell_border

            current_row += 1

        # --- 6. FOOTER SIGNATURES & FOOTER BLOCK ---
        footer_row = current_row + 3
        ws.merge_cells(f"B{footer_row}:D{footer_row}")
        ws.cell(row=footer_row, column=2, value="អ្នករៀបចំ ៖ ____________________").font = font_footer
        ws.cell(row=footer_row, column=2).alignment = align_left

        ws.merge_cells(f"G{footer_row}:I{footer_row}")
        ws.cell(row=footer_row, column=7, value="អ្នកអនុម័ត ៖ ____________________").font = font_footer
        ws.cell(row=footer_row, column=7).alignment = align_right
        
        ws.merge_cells(f"A{footer_row+2}:I{footer_row+2}")
        ws.cell(row=footer_row + 2, column=1, value="PNK  Shop - របាយការណ៍ផ្លូវការ ").font = font_footer
        ws.cell(row=footer_row + 2, column=1).alignment = align_center

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 28
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 42
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 26

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="PNK_Orders_Report_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response
    else:
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="PNK_Orders_Report_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
        response.write('\ufeff')

        writer = csv.writer(response)
        writer.writerow(['ល.រ', 'ORDER ID', 'អតិថិជន ', 'លេខទូរស័ព្ទ', 'មុខទំនិញ ', 'ទូទាត់', 'ស្ថានភាព', 'សរុប ($)', 'កាលបរិច្ឆេទ'])

        for idx, ord_obj in enumerate(orders_qs, 1):
            items_str = ", ".join([f"{item.product.name} (x{item.quantity})" for item in ord_obj.items.all() if item.product]) or "-"
            cust_name = f"{ord_obj.last_name} {ord_obj.first_name}".strip() or (ord_obj.user.username if ord_obj.user else "Guest")
            writer.writerow([
                idx,
                f"#{ord_obj.id}",
                cust_name,
                ord_obj.phone or "",
                items_str,
                ord_obj.payment_method or "COD",
                ord_obj.status,
                f"{ord_obj.total_amount:.2f}",
                ord_obj.created_at.strftime("%d/%m/%Y %H:%M")
            ])

        return response

@login_required
@user_passes_test(is_admin)
def update_order_status(request, order_id):
    if request.method == 'POST':
        order = get_object_or_404(Order, id=order_id)
        old_status = order.status
        new_status = request.POST.get('status')
        admin_note = request.POST.get('admin_note', '').strip()
        if new_status in ['Pending', 'Processing', 'Completed', 'Rejected']:
            # 🎯 Inventory Stock Management Integration
            if old_status != 'Rejected' and new_status == 'Rejected':
                # Restore stock when order is rejected
                for item in order.items.all():
                    if item.product:
                        item.product.stock += item.quantity
                        item.product.save()
            elif old_status == 'Rejected' and new_status in ['Completed', 'Pending', 'Processing']:
                # Re-deduct stock if previously rejected order is un-rejected
                for item in order.items.all():
                    if item.product:
                        item.product.stock = max(0, item.product.stock - item.quantity)
                        item.product.save()

            order.status = new_status
            if admin_note:
                order.admin_note = admin_note
            order.save()

            # Recalculate KPI summary stats for real-time live AJAX updates
            completed_orders = Order.objects.filter(status__iexact='Completed')
            rejected_orders = Order.objects.filter(status__iexact='Rejected')
            pending_orders = Order.objects.exclude(status__in=['Completed', 'Rejected', 'completed', 'rejected'])
            total_revenue = completed_orders.aggregate(total=Sum('total_amount'))['total'] or 0

            return JsonResponse({
                'success': True,
                'new_status': order.status,
                'admin_note': order.admin_note,
                'count_pending': pending_orders.count(),
                'count_completed': completed_orders.count(),
                'count_rejected': rejected_orders.count(),
                'total_revenue': float(total_revenue),
            })
    return JsonResponse({'success': False})

@login_required
@user_passes_test(is_admin)
def confirm_order(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    order.status = 'Completed'
    order.save()

    completed_orders = Order.objects.filter(status__iexact='Completed')
    rejected_orders = Order.objects.filter(status__iexact='Rejected')
    pending_orders = Order.objects.exclude(status__in=['Completed', 'Rejected', 'completed', 'rejected'])
    total_revenue = completed_orders.aggregate(total=Sum('total_amount'))['total'] or 0

    return JsonResponse({
        'status': 'success',
        'new_status': order.status,
        'count_pending': pending_orders.count(),
        'count_completed': completed_orders.count(),
        'count_rejected': rejected_orders.count(),
        'total_revenue': float(total_revenue),
    })

# ==============================================================================
# 5. USER & CUSTOMER MANAGEMENT
# ==============================================================================
@login_required
@user_passes_test(is_admin)
def manage_users(request):
    from django.db.models import Count, Q
    search_query = request.GET.get('search', '').strip()
    role_filter = request.GET.get('role', 'all')
    status_filter = request.GET.get('status', 'all')
    sort_by = request.GET.get('sort', 'newest')

    users_qs = User.objects.select_related('profile').annotate(order_count=Count('order')).all()

    # Mark new customers as seen/read in session
    latest_cust = User.objects.filter(is_staff=False, is_superuser=False).order_by('-id').first()
    if latest_cust:
        request.session['last_seen_customer_id'] = latest_cust.id
        request.session.modified = True

    # Calculate statistics for KPI Cards before filtering
    total_users = users_qs.count()
    active_users = users_qs.filter(is_active=True).count()
    disabled_users = users_qs.filter(is_active=False).count()
    admin_users = users_qs.filter(Q(is_superuser=True) | Q(is_staff=True)).count()

    # Search query
    if search_query:
        users_qs = users_qs.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(profile__phone__icontains=search_query)
        )

    # Role filter
    if role_filter == 'admin':
        users_qs = users_qs.filter(Q(is_superuser=True) | Q(is_staff=True))
    elif role_filter == 'customer':
        users_qs = users_qs.filter(is_superuser=False, is_staff=False)

    # Status filter
    if status_filter == 'active':
        users_qs = users_qs.filter(is_active=True)
    elif status_filter == 'disabled':
        users_qs = users_qs.filter(is_active=False)

    # Sorting
    if sort_by == 'oldest':
        users_qs = users_qs.order_by('date_joined')
    elif sort_by == 'orders':
        users_qs = users_qs.order_by('-order_count', '-date_joined')
    elif sort_by == 'name_asc':
        users_qs = users_qs.order_by('last_name', 'first_name', 'username')
    else:
        users_qs = users_qs.order_by('-date_joined')

    context = {
        'users': users_qs,
        'total_users': total_users,
        'active_users': active_users,
        'disabled_users': disabled_users,
        'admin_users': admin_users,
        'search_query': search_query,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'sort_by': sort_by,
    }
    return render(request, 'accounts/dashboard/manage_users.html', context)

@login_required
@user_passes_test(is_admin)
def add_user(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        role = request.POST.get('role', 'customer')

        if not username or not password:
            messages.error(request, "សូមបំពេញ ឈ្មោះអ្នកប្រើប្រាស់ និង ពាក្យសម្ងាត់!")
        elif User.objects.filter(username=username).exists():
            messages.error(request, f"ឈ្មោះអ្នកប្រើប្រាស់ '{username}' មានរួចហើយ!")
        else:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )
            if role == 'admin':
                user.is_staff = True
                user.is_superuser = True
                user.save()
            if hasattr(user, 'profile'):
                user.profile.phone = phone
                user.profile.save()
            messages.success(request, f"បានបង្កើតគណនី '{username}' ដោយជោគជ័យ!")
    return redirect('accounts:manage_users')

@login_required
@user_passes_test(is_admin)
def user_detail_api(request, user_id):
    user_obj = get_object_or_404(User.objects.select_related('profile'), id=user_id)
    orders = user_obj.order_set.all().order_by('-created_at')[:5]
    orders_data = [{
        'id': o.id,
        'total_amount': float(o.total_amount),
        'status': o.status,
        'created_at': o.created_at.strftime('%Y-%m-%d %H:%M')
    } for o in orders]

    profile_img_url = user_obj.profile.image.url if (hasattr(user_obj, 'profile') and user_obj.profile.image) else None
    phone = user_obj.profile.phone if hasattr(user_obj, 'profile') else "-"

    data = {
        'id': user_obj.id,
        'username': user_obj.username,
        'first_name': user_obj.first_name,
        'last_name': user_obj.last_name,
        'email': user_obj.email or "-",
        'phone': phone or "-",
        'is_active': user_obj.is_active,
        'is_superuser': user_obj.is_superuser,
        'date_joined': user_obj.date_joined.strftime('%Y-%m-%d %H:%M'),
        'profile_image': profile_img_url,
        'total_orders': user_obj.order_set.count(),
        'recent_orders': orders_data
    }
    return JsonResponse({'success': True, 'user': data})

@login_required
@user_passes_test(is_admin)
def toggle_user_active(request, user_id):
    user_obj = get_object_or_404(User, id=user_id)
    if user_obj.is_superuser:
        return JsonResponse({'success': False, 'error': 'មិនអាចបិទគណនី Admin បានទេ!'})
    user_obj.is_active = not user_obj.is_active
    user_obj.save()
    return JsonResponse({'success': True, 'is_active': user_obj.is_active})

@login_required
@user_passes_test(is_admin)
def delete_user(request, user_id):
    user_to_delete = get_object_or_404(User, id=user_id)
    if user_to_delete.is_superuser:
        return JsonResponse({'success': False, 'error': 'មិនអាចលុបគណនី Admin បានទេ!'})

    if request.method == 'POST':
        user_to_delete.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'សំណើមិនត្រឹមត្រូវ'})

# ==============================================================================
# 6. SLIDESHOW MANAGEMENT
# ==============================================================================
# 6. BLOG MANAGEMENT MODULE
# ==============================================================================
@login_required
@user_passes_test(is_admin)
def manage_blog(request):
    from shop.models import Blog
    search_query = request.GET.get('search', '').strip()
    
    blogs_qs = Blog.objects.all().order_by('-created_at')
    if search_query:
        blogs_qs = blogs_qs.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(slug__icontains=search_query)
        )

    context = {
        'blogs': blogs_qs,
        'search_query': search_query,
        'total_blogs': Blog.objects.count(),
        'blogs_with_image': Blog.objects.exclude(image='').count(),
    }
    return render(request, 'accounts/dashboard/manage_blog.html', context)

@login_required
@user_passes_test(is_admin)
def add_blog(request):
    from shop.models import Blog
    from django.utils.text import slugify
    import time

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        custom_slug = request.POST.get('slug', '').strip()
        image = request.FILES.get('image')

        if name and description:
            slug = slugify(custom_slug if custom_slug else name)
            if not slug:
                slug = f"blog-{int(time.time())}"
            
            unique_slug = slug
            counter = 1
            while Blog.objects.filter(slug=unique_slug).exists():
                unique_slug = f"{slug}-{counter}"
                counter += 1

            Blog.objects.create(
                name=name,
                slug=unique_slug,
                description=description,
                image=image
            )
            log_activity(request.user, f"បន្ថែម Blog ថ្មី [{name}]", f"Slug: {unique_slug}", "bi-journal-plus", "text-success")
            messages.success(request, f"បានបន្ថែមអត្ថបទថ្មី [{name}] ដោយជោគជ័យ!")
        else:
            messages.error(request, "សូមបំពេញចំណងជើង និង ការពិពណ៌នាឱ្យបានត្រឹមត្រូវ!")
    return redirect('accounts:manage_blog')

@login_required
@user_passes_test(is_admin)
def edit_blog(request, pk):
    from shop.models import Blog
    from django.utils.text import slugify

    blog = get_object_or_404(Blog, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        custom_slug = request.POST.get('slug', '').strip()
        image = request.FILES.get('image')

        if name and description:
            blog.name = name
            blog.description = description
            if image:
                blog.image = image

            slug_source = custom_slug if custom_slug else name
            new_slug = slugify(slug_source)
            if not new_slug:
                new_slug = blog.slug or f"blog-{blog.id}"

            if new_slug != blog.slug:
                unique_slug = new_slug
                counter = 1
                while Blog.objects.filter(slug=unique_slug).exclude(pk=pk).exists():
                    unique_slug = f"{new_slug}-{counter}"
                    counter += 1
                blog.slug = unique_slug

            blog.save()
            log_activity(request.user, f"កែប្រែ Blog [{blog.name}]", f"ID #{blog.id}", "bi-pencil-square", "text-warning")
            messages.success(request, f"បានកែប្រែអត្ថបទ [{blog.name}] ដោយជោគជ័យ!")
        else:
            messages.error(request, "សូមបំពេញព័ត៌មានចាំបាច់ឱ្យបានត្រឹមត្រូវ!")
    return redirect('accounts:manage_blog')


@login_required
@user_passes_test(is_admin)
def delete_blog(request, pk):
    from shop.models import Blog
    blog = get_object_or_404(Blog, pk=pk)
    blog_title = blog.name
    blog.delete()
    messages.success(request, f"បានលុបអត្ថបទ [{blog_title}] រួចរាល់!")
    return redirect('accounts:manage_blog')

@login_required
@user_passes_test(is_admin)
def dashboard_blog_detail(request, pk):
    from shop.models import Blog
    blog = get_object_or_404(Blog, pk=pk)
    recent_blogs = Blog.objects.exclude(pk=pk).order_by('-created_at')[:4]
    context = {
        'blog': blog,
        'recent_blogs': recent_blogs,
    }
    return render(request, 'accounts/dashboard/blog-details.html', context)


@login_required
@user_passes_test(is_admin)
def manage_slideshow(request):
    slides = Slideshow.objects.all().order_by('-id')
    return render(request, 'accounts/dashboard/manage_slideshow.html', {'slides': slides})

@login_required
@user_passes_test(is_admin)
def add_slideshow(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        image = request.FILES.get('image')
        link = request.POST.get('link')
        description = request.POST.get('description')
        title_color = request.POST.get('title_color', '#ffffff')
        desc_color = request.POST.get('desc_color', '#e0e0e0')
        shadow_color = request.POST.get('shadow_color', '#000000')

        if title and image:
            Slideshow.objects.create(
                title=title,
                image=image,
                link=link,
                description=description,
                title_color=title_color,
                desc_color=desc_color,
                shadow_color=shadow_color
            )
            messages.success(request, "បានបន្ថែមរូបភាព Slideshow ថ្មីដោយជោគជ័យ!")
        else:
            messages.error(request, "សូមបំពេញព័ត៌មានចាំបាច់ (ចំណងជើង និងរូបភាព) ឱ្យបានត្រឹមត្រូវ!")
    return redirect('accounts:manage_slideshow')

@login_required
@user_passes_test(is_admin)
def edit_slideshow(request, pk):
    slide = get_object_or_404(Slideshow, pk=pk)
    if request.method == 'POST':
        slide.title = request.POST.get('title')
        slide.link = request.POST.get('link')
        slide.description = request.POST.get('description')
        slide.title_color = request.POST.get('title_color')
        slide.desc_color = request.POST.get('desc_color')
        slide.shadow_color = request.POST.get('shadow_color')
        
        if request.FILES.get('image'):
            slide.image = request.FILES.get('image')
            
        slide.save()
        messages.success(request, "បានកែប្រែជោគជ័យ!")
    return redirect('accounts:manage_slideshow')

@login_required
@user_passes_test(is_admin)
def delete_slideshow(request, pk):
    slide = get_object_or_404(Slideshow, pk=pk)
    slide.delete()
    messages.success(request, "បានលុបរូបភាព Slideshow នេះចេញពីប្រព័ន្ធហើយ!")
    return redirect('accounts:manage_slideshow')

# ==============================================================================
# 7. REPORTS & SETTINGS
# ==============================================================================
@login_required
@user_passes_test(is_admin)
def report_page(request):
    from django.core.paginator import Paginator
    from shop.models import Brand, Product, Order

    orders_qs, filters = get_filtered_orders(request)

    # Aggregate over full filtered QuerySet
    completed_orders = orders_qs.filter(status__iexact='Completed')
    rejected_orders = orders_qs.filter(status__iexact='Rejected')
    pending_orders = orders_qs.exclude(status__iexact='Completed').exclude(status__iexact='Rejected')

    total_sales = completed_orders.aggregate(total=Sum('total_amount'))['total'] or 0
    total_orders_count = orders_qs.count()
    count_completed = completed_orders.count()
    count_rejected = rejected_orders.count()
    count_pending = pending_orders.count()

    # Percentages
    completion_rate = round((count_completed / total_orders_count * 100), 1) if total_orders_count > 0 else 0
    rejection_rate = round((count_rejected / total_orders_count * 100), 1) if total_orders_count > 0 else 0

    # Top VIP Buyers (completed orders)
    top_customers = (
        Order.objects.filter(status__iexact='Completed')
        .values('first_name', 'last_name', 'phone', 'email', 'user__username')
        .annotate(total_spent=Sum('total_amount'), total_orders=Count('id'))
        .order_by('-total_spent')[:5]
    )

    # Low Stock Alert Products (stock <= 5)
    low_stock_products = Product.objects.filter(stock__lte=5).order_by('stock')[:6]

    # Top selling products
    top_products = Product.objects.annotate(
        sales_count=Count('orderitem')
    ).filter(sales_count__gt=0).order_by('-sales_count')[:10]

    all_brands = Brand.objects.all().order_by('name')
    all_products = Product.objects.all().order_by('name')

    # Pagination: 15 orders per page
    paginator = Paginator(orders_qs, 15)
    page_number = request.GET.get('page', 1)
    orders_page = paginator.get_page(page_number)

    # Calculate page total sum
    page_total_amount = sum(o.total_amount for o in orders_page if o.total_amount)

    context = {
        'orders': orders_page,
        'total_sales': total_sales,
        'total_orders_count': total_orders_count,
        'count_pending': count_pending,
        'count_completed': count_completed,
        'count_rejected': count_rejected,
        'completion_rate': completion_rate,
        'rejection_rate': rejection_rate,
        'page_total_amount': page_total_amount,
        'top_customers': top_customers,
        'low_stock_products': low_stock_products,
        'top_products': top_products,
        'all_brands': all_brands,
        'all_products': all_products,
        **filters
    }
    return render(request, 'accounts/dashboard/report.html', context)

@login_required
@user_passes_test(is_admin)
def export_top_products_csv(request):
    import csv
    from shop.models import Product
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="top_selling_products.csv"'

    writer = csv.writer(response)
    writer.writerow(['Product Name', 'Category', 'Brand', 'Price ($)', 'Total Units Sold'])

    top_products = Product.objects.annotate(
        sales_count=Count('orderitem')
    ).filter(sales_count__gt=0).order_by('-sales_count')

    for p in top_products:
        writer.writerow([
            p.name,
            p.category.name if p.category else '-',
            p.brand.name if p.brand else '-',
            p.price,
            p.sales_count
        ])
    log_activity(request.user, "ទាញយក CSV ទូរស័ព្ទលក់ដាច់", "Exported Top Products CSV", "bi-file-earmark-spreadsheet", "text-success")
    return response

@login_required
@user_passes_test(is_admin)
def export_vip_customers_csv(request):
    import csv
    orders_qs, filters = get_filtered_orders(request)

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    filename = f"vip_customers_report_{filters['date_filter']}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(['Customer Name', 'Username', 'Phone', 'Email', 'Total Orders', 'Total Spent ($)'])

    top_customers = (
        orders_qs.filter(status__iexact='Completed')
        .values('first_name', 'last_name', 'phone', 'email', 'user__username')
        .annotate(total_spent=Sum('total_amount'), total_orders=Count('id'))
        .order_by('-total_spent')
    )

    for c in top_customers:
        writer.writerow([
            f"{c['last_name']} {c['first_name']}".strip() or c['user__username'],
            c['user__username'],
            c['phone'] or '-',
            c['email'] or '-',
            c['total_orders'],
            c['total_spent']
        ])
    log_activity(request.user, "ទាញយក CSV អតិថិជន VIP", "Exported VIP Customers CSV", "bi-file-earmark-spreadsheet", "text-success")
    return response

@login_required
@user_passes_test(is_admin)
def export_low_stock_csv(request):
    import csv
    from shop.models import Product
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="low_stock_inventory.csv"'

    writer = csv.writer(response)
    writer.writerow(['Product Name', 'Category', 'Brand', 'Price ($)', 'Current Stock', 'Status'])

    low_stock_products = Product.objects.filter(stock__lte=5).order_by('stock')

    for p in low_stock_products:
        status_str = "Out of Stock" if p.stock == 0 else "Low Stock"
        writer.writerow([
            p.name,
            p.category.name if p.category else '-',
            p.brand.name if p.brand else '-',
            p.price,
            p.stock,
            status_str
        ])
    log_activity(request.user, "ទាញយក CSV របាយការណ៍ស្តុក", "Exported Low Stock CSV", "bi-file-earmark-spreadsheet", "text-warning")
    return response

@login_required
@user_passes_test(is_admin)
def upload_page(request):
    return render(request, 'accounts/dashboard/upload.html')

@login_required
@user_passes_test(is_admin)
def settings_page(request):
    store_settings = StoreSetting.get_settings()

    if request.method == 'POST':
        store_name = request.POST.get('store_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        email = request.POST.get('email', '').strip()
        currency = request.POST.get('currency', 'USD').strip()
        address = request.POST.get('address', '').strip()
        facebook_link = request.POST.get('facebook_link', '').strip()
        telegram_link = request.POST.get('telegram_link', '').strip()
        map_iframe = request.POST.get('map_iframe', '').strip()
        chat_auto_reply = request.POST.get('chat_auto_reply', '').strip()

        if store_name:
            store_settings.store_name = store_name
        if phone:
            store_settings.phone = phone
        if email:
            store_settings.email = email
        if currency:
            store_settings.currency = currency
        if address:
            store_settings.address = address
        store_settings.facebook_link = facebook_link
        store_settings.telegram_link = telegram_link
        store_settings.map_iframe = map_iframe
        store_settings.enable_auto_reply = 'enable_auto_reply' in request.POST
        if chat_auto_reply:
            store_settings.chat_auto_reply = chat_auto_reply

        # Handle logo upload
        if 'logo' in request.FILES:
            store_settings.logo = request.FILES['logo']

        store_settings.save()

        log_activity(request.user, "បច្ចុប្បន្នភាពការកំណត់ហាង", f"Updated settings for {store_settings.store_name}", "bi-gear-fill", "text-success")
        messages.success(request, "បានរក្សាទុកការកំណត់ហាងដោយជោគជ័យ!")
        return redirect('accounts:settings_page')

    return render(request, 'accounts/dashboard/settings.html', {'store_settings': store_settings})

# ==============================================================================
# 8. AUTHENTICATION & PROFILE
# ==============================================================================
def login_view(request):
    if request.method == 'POST':
        login_input = request.POST.get('login', '').strip() or request.POST.get('login_input', '').strip()
        password = request.POST.get('password', '')

        try:
            user_obj = User.objects.get(
                Q(username=login_input) | 
                Q(email=login_input) | 
                Q(profile__phone=login_input)
            )
            real_username = user_obj.username
        except User.DoesNotExist:
            real_username = None

        user = authenticate(request, username=real_username, password=password)

        if user is not None:
            auth_login(request, user)

            # Dynamic Session Expiry Control
            if user.is_superuser or user.is_staff:
                request.session.set_expiry(1800)      # 30 នាទី (Admin / Staff សុវត្ថិភាពខ្ពស់)
            else:
                request.session.set_expiry(0)         # ផ្ដាច់ Session ភ្លាមៗពេលបិទ Browser សម្រាប់ Customer

            messages.success(request, f"ស្វាគមន៍មកកាន់ PNK SHOP, {user.first_name if user.first_name else user.username}!")
            return redirect('core:home')
        else:
            messages.error(request, "ឈ្មោះ ឬ លេខសម្ងាត់មិនត្រឹមត្រូវ!")
            return redirect('account_login')
            
    return render(request, 'account/login.html')

def logout_view(request):
    is_expired = request.GET.get('session_expired') == '1'
    logout(request)
    if is_expired:
        messages.warning(request, "Session របស់អ្នកបានផុតកំណត់ដោយសារគ្មានសកម្មភាពលើសពី ៣០ នាទី។ សូម Login ឡើងវិញ!")
    return redirect('account_login')

def signup_view(request):
    if request.method == "POST":
        form = SignupForm(request.POST)
        if form.is_valid():
            user = form.save() 
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            request.session.set_expiry(0)             # ផ្ដាច់ Session ភ្លាមៗពេលបិទ Browser សម្រាប់ Customer ចុះឈ្មោះថ្មី
            return JsonResponse({'success': True, 'message': 'ចុះឈ្មោះជោគជ័យ'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)
            
    return render(request, 'account/signup.html')

@login_required
def profile_view(request):
    user = request.user
    profile, created = UserProfile.objects.get_or_create(user=user)
    
    if request.method == 'POST':
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.email = request.POST.get('email', user.email)
        user.save()
        
        profile.phone = request.POST.get('phone', profile.phone)
        if request.FILES.get('image'):
            profile.image = request.FILES.get('image')
        profile.save()
        profile.refresh_from_db() 
        
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'success',
                'message': 'បានផ្លាស់ប្តូរព័ត៌មានគណនីដោយជោគជ័យ!',
                'new_image_url': profile.image.url if profile.image else ''
            })
        messages.success(request, 'បានផ្លាស់ប្តូរព័ត៌មានគណនីដោយជោគជ័យ!')
        return redirect('accounts:profile')
            
    return render(request, 'accounts/shop/profile.html', {'user': user, 'profile': profile})


@login_required
@user_passes_test(is_admin)
@xframe_options_exempt
def export_orders_pdf(request):
    orders_qs, filters = get_filtered_orders(request)

    count_all = orders_qs.count()
    count_completed = orders_qs.filter(status__iexact='Completed').count()
    count_pending = orders_qs.exclude(status__iexact='Completed').exclude(status__iexact='Rejected').count()
    count_rejected = orders_qs.filter(status__iexact='Rejected').count()
    total_revenue = orders_qs.filter(status__iexact='Completed').aggregate(total=Sum('total_amount'))['total'] or 0

    context = {
        'orders': orders_qs,
        'count_all': count_all,
        'count_completed': count_completed,
        'count_pending': count_pending,
        'count_rejected': count_rejected,
        'total_revenue': total_revenue,
        'period_label': filters['period_label'],
        'status_filter': filters['status_filter'],
        'generated_at': timezone.now(),
    }
    return render(request, 'accounts/dashboard/order_report_pdf.html', context)

@login_required
@user_passes_test(is_admin)
@xframe_options_exempt
def export_top_products_pdf(request):
    from shop.models import OrderItem
    top_products_qs = (
        OrderItem.objects.filter(order__status__iexact='Completed')
        .values('product__id', 'product__name', 'product__brand__name', 'product__price')
        .annotate(
            total_qty=Sum('quantity'),
            total_revenue=Sum(F('quantity') * F('price'))
        )
        .order_by('-total_qty')
    )

    total_items_sold = sum(p['total_qty'] or 0 for p in top_products_qs)
    total_revenue = sum(p['total_revenue'] or 0 for p in top_products_qs)

    context = {
        'top_products': top_products_qs,
        'total_top_products': len(top_products_qs),
        'total_items_sold': total_items_sold,
        'total_revenue': total_revenue,
        'generated_at': timezone.now(),
    }
    return render(request, 'accounts/dashboard/top_products_report_pdf.html', context)

@login_required
@user_passes_test(is_admin)
@xframe_options_exempt
def export_vip_customers_pdf(request):
    vip_qs = (
        Order.objects.filter(status__iexact='Completed')
        .values('first_name', 'last_name', 'email', 'phone', 'user__username')
        .annotate(
            total_spent=Sum('total_amount'),
            order_count=Count('id'),
            last_order=Max('created_at')
        )
        .order_by('-total_spent')
    )

    total_vips = len(vip_qs)
    total_vip_orders = sum(v['order_count'] or 0 for v in vip_qs)
    total_vip_spend = sum(v['total_spent'] or 0 for v in vip_qs)

    context = {
        'vip_customers': vip_qs,
        'total_vips': total_vips,
        'total_vip_orders': total_vip_orders,
        'total_vip_spend': total_vip_spend,
        'generated_at': timezone.now(),
    }
    return render(request, 'accounts/dashboard/vip_customers_report_pdf.html', context)

@login_required
@user_passes_test(is_admin)
@xframe_options_exempt
def export_low_stock_pdf(request):
    low_stock_qs = Product.objects.filter(stock__lt=5).order_by('stock')

    count_low = low_stock_qs.filter(stock__gt=0).count()
    count_zero = low_stock_qs.filter(stock=0).count()
    stock_value = sum(p.stock * float(p.price) for p in low_stock_qs)

    context = {
        'low_stock_products': low_stock_qs,
        'count_low': count_low,
        'count_zero': count_zero,
        'stock_value': stock_value,
        'generated_at': timezone.now(),
    }
    return render(request, 'accounts/dashboard/low_stock_report_pdf.html', context)

@login_required
@user_passes_test(is_admin)
def export_top_products_csv(request):
    from shop.models import OrderItem
    top_products_qs = (
        OrderItem.objects.filter(order__status__iexact='Completed')
        .values('product__id', 'product__name', 'product__brand__name', 'product__price')
        .annotate(
            total_qty=Sum('quantity'),
            total_revenue=Sum(F('quantity') * F('price'))
        )
        .order_by('-total_qty')
    )

    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Top Products Report"
        ws.views.sheetView[0].showGridLines = True

        # Page Setup (Landscape A4 + Fit All Columns on 1 Page Width)
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.6
        ws.page_margins.bottom = 0.6
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

        ws.oddHeader.left.text = "PNK SHOP - របាយការណ៍ទូរស័ព្ទលក់ដាច់"
        ws.oddHeader.right.text = "កាលបរិច្ឆេទ: &[Date]"
        ws.oddFooter.left.text = "PNK SHOP"
        ws.oddFooter.center.text = "ទំព័រ &[Page] នៃ &[Pages]"

        font_title = Font(name="Khmer OS Muol Light", size=15, bold=True, color="FFFFFF")
        font_subtitle = Font(name="Khmer OS Battambang", size=11, bold=True, color="E2E8F0")
        font_meta_lbl = Font(name="Khmer OS Siemreap", size=10, bold=True, color="334155")
        font_meta_val = Font(name="Khmer OS Siemreap", size=10, color="0F172A")
        font_kpi_lbl = Font(name="Khmer OS Battambang", size=9, bold=True, color="475569")
        font_tbl_header = Font(name="Khmer OS Battambang", size=10.5, bold=True, color="FFFFFF")
        font_data = Font(name="Khmer OS Siemreap", size=10, color="0F172A")
        font_footer = Font(name="Khmer OS Siemreap", size=9.5, italic=True, color="475569")

        fill_header_banner = PatternFill(start_color="198754", end_color="146C43", fill_type="solid")
        fill_tbl_header = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_kpi_card = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        thin_border = Side(style='thin', color='CBD5E1')
        cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

        ws.merge_cells("A1:G1")
        ws["A1"] = "PNK SHOP"
        ws["A1"].font = font_title
        ws["A1"].alignment = align_center

        ws.merge_cells("A2:G2")
        ws["A2"] = "របាយការណ៍ទូរស័ព្ទលក់ដាច់បំផុត (Top Selling Products Report)"
        ws["A2"].font = font_subtitle
        ws["A2"].alignment = align_center

        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 26
        for r in range(1, 3):
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = fill_header_banner

        now_str = timezone.now().strftime('%d-%m-%Y %I:%M %p')
        ws.merge_cells("A4:E4")
        ws["A4"] = "📍 អាសយដ្ឋាន: ភូមិនិគមន៍លើ ខេត្តត្បូងឃ្មុំ, ព្រះរាជាណាចក្រកម្ពុជា"
        ws["A4"].font = font_meta_lbl

        ws.merge_cells("A5:E5")
        ws["A5"] = "📞 ទូរស័ព្ទ: 096 29 647 13 | ✉️ អ៊ីមែល: dana267yue@gmail.com"
        ws["A5"].font = font_meta_lbl

        ws.merge_cells("A6:E6")
        ws["A6"] = f"📅 កាលបរិច្ឆេទបង្កើត: {now_str}"
        ws["A6"].font = font_meta_lbl

        ws.merge_cells("F4:G4")
        ws["F4"] = "អ្នកបង្កើត: Administrator"
        ws["F4"].font = font_meta_lbl
        ws["F4"].alignment = align_right

        total_items_sold = sum(p['total_qty'] or 0 for p in top_products_qs)
        total_revenue = sum(p['total_revenue'] or 0 for p in top_products_qs)

        kpi_cards = [
            ("A8:B8", "A9:B9", "A8", "A9", "មុខទំនិញលក់ដាច់សរុប", len(top_products_qs), "0F172A"),
            ("C8:D8", "C9:D9", "C8", "C9", "ចំនួនលក់សរុប (គ្រឿង)", total_items_sold, "16A34A"),
            ("E8:G8", "E9:G9", "E8", "E9", "ចំណូលសរុប ($)", f"${total_revenue:.2f}", "16A34A")
        ]

        ws.row_dimensions[8].height = 20
        ws.row_dimensions[9].height = 28

        for lbl_range, val_range, lbl_cell, val_cell, label_text, val_text, color_hex in kpi_cards:
            if ":" in lbl_range:
                ws.merge_cells(lbl_range)
                ws.merge_cells(val_range)
            ws[lbl_cell] = label_text
            ws[lbl_cell].font = font_kpi_lbl
            ws[lbl_cell].alignment = align_center
            ws[val_cell] = val_text
            ws[val_cell].font = Font(name="Khmer OS Siemreap", size=13, bold=True, color=color_hex)
            ws[val_cell].alignment = align_center

        for r in range(8, 10):
            for c in range(1, 8):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_kpi_card
                cell.border = cell_border

        headers = ["ល.រ", "ID ទំនិញ", "ឈ្មោះទូរស័ព្ទ/ផលិតផល", "ម៉ាក (Brand)", "ចំនួនលក់ (គ្រឿង)", "តម្លៃរាយ ($)", "ចំណូលសរុប ($)"]
        ws.row_dimensions[11].height = 34
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col_num, value=header)
            cell.font = font_tbl_header
            cell.fill = fill_tbl_header
            cell.alignment = align_center
            cell.border = cell_border

        current_row = 12
        for idx, item in enumerate(top_products_qs, 1):
            pname = str(item['product__name'] or "")
            if len(pname) > 40:
                ws.row_dimensions[current_row].height = 42
            elif len(pname) > 22:
                ws.row_dimensions[current_row].height = 32
            else:
                ws.row_dimensions[current_row].height = 26

            c1 = ws.cell(row=current_row, column=1, value=idx)
            c1.alignment = align_center

            c2 = ws.cell(row=current_row, column=2, value=f"#{item['product__id']}")
            c2.alignment = align_center

            c3 = ws.cell(row=current_row, column=3, value=item['product__name'])
            c3.alignment = align_left

            c4 = ws.cell(row=current_row, column=4, value=item['product__brand__name'] or "-")
            c4.alignment = align_center

            c5 = ws.cell(row=current_row, column=5, value=item['total_qty'])
            c5.alignment = align_center

            c6 = ws.cell(row=current_row, column=6, value=float(item['product__price']))
            c6.number_format = "$#,##0.00"
            c6.alignment = align_right

            c7 = ws.cell(row=current_row, column=7, value=float(item['total_revenue'] or 0))
            c7.number_format = "$#,##0.00"
            c7.alignment = align_right

            row_fill = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
            for c in range(1, 8):
                cell = ws.cell(row=current_row, column=c)
                if row_fill.fill_type: cell.fill = row_fill
                cell.border = cell_border
                if c not in [6, 7]: cell.font = font_data
                else: cell.font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="16A34A")

            current_row += 1

        footer_row = current_row + 3
        ws.merge_cells(f"B{footer_row}:C{footer_row}")
        ws.cell(row=footer_row, column=2, value="អ្នករៀបចំ ៖ ____________________").font = font_footer
        ws.cell(row=footer_row, column=2).alignment = align_left

        ws.merge_cells(f"E{footer_row}:G{footer_row}")
        ws.cell(row=footer_row, column=5, value="អ្នកអនុម័ត ៖ ____________________").font = font_footer
        ws.cell(row=footer_row, column=5).alignment = align_right

        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 38
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 18

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="PNK_Top_Products_Report_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response
    else:
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="PNK_Top_Products_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
        response.write('\ufeff')
        writer = csv.writer(response)
        writer.writerow(["ល.រ", "ID ទំនិញ", "ឈ្មោះទូរស័ព្ទ/ផលិតផល", "ម៉ាក (Brand)", "ចំនួនលក់ (គ្រឿង)", "តម្លៃរាយ ($)", "ចំណូលសរុប ($)"])
        for idx, item in enumerate(top_products_qs, 1):
            writer.writerow([idx, f"#{item['product__id']}", item['product__name'], item['product__brand__name'] or "-", item['total_qty'], f"{item['product__price']:.2f}", f"{item['total_revenue']:.2f}"])
        return response

@login_required
@user_passes_test(is_admin)
def export_vip_customers_csv(request):
    vip_qs = (
        Order.objects.filter(status__iexact='Completed')
        .values('first_name', 'last_name', 'email', 'phone', 'user__username')
        .annotate(
            total_spent=Sum('total_amount'),
            order_count=Count('id'),
            last_order=Max('created_at')
        )
        .order_by('-total_spent')
    )

    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VIP Customers Report"
        ws.views.sheetView[0].showGridLines = True

        # Page Setup (Landscape A4 + Fit All Columns on 1 Page Width)
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.6
        ws.page_margins.bottom = 0.6
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

        ws.oddHeader.left.text = "PNK SHOP - របាយការណ៍អតិថិជន VIP"
        ws.oddHeader.right.text = "កាលបរិច្ឆេទ: &[Date]"
        ws.oddFooter.left.text = "PNK SHOP"
        ws.oddFooter.center.text = "ទំព័រ &[Page] នៃ &[Pages]"

        font_title = Font(name="Khmer OS Muol Light", size=15, bold=True, color="FFFFFF")
        font_subtitle = Font(name="Khmer OS Battambang", size=11, bold=True, color="E2E8F0")
        font_meta_lbl = Font(name="Khmer OS Siemreap", size=10, bold=True, color="334155")
        font_meta_val = Font(name="Khmer OS Siemreap", size=10, color="0F172A")
        font_kpi_lbl = Font(name="Khmer OS Battambang", size=9, bold=True, color="475569")
        font_tbl_header = Font(name="Khmer OS Battambang", size=10.5, bold=True, color="FFFFFF")
        font_data = Font(name="Khmer OS Siemreap", size=10, color="0F172A")
        font_footer = Font(name="Khmer OS Siemreap", size=9.5, italic=True, color="475569")

        fill_header_banner = PatternFill(start_color="198754", end_color="146C43", fill_type="solid")
        fill_tbl_header = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_kpi_card = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        thin_border = Side(style='thin', color='CBD5E1')
        cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

        ws.merge_cells("A1:G1")
        ws["A1"] = "PNK SHOP"
        ws["A1"].font = font_title
        ws["A1"].alignment = align_center

        ws.merge_cells("A2:G2")
        ws["A2"] = "របាយការណ៍អតិថិជន VIP (VIP Customers Report)"
        ws["A2"].font = font_subtitle
        ws["A2"].alignment = align_center

        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 26
        for r in range(1, 3):
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = fill_header_banner

        now_str = timezone.now().strftime('%d-%m-%Y %I:%M %p')
        ws.merge_cells("A4:E4")
        ws["A4"] = "📍 អាសយដ្ឋាន: ភូមិនិគមន៍លើ ខេត្តត្បូងឃ្មុំ, ព្រះរាជាណាចក្រកម្ពុជា"
        ws["A4"].font = font_meta_lbl

        ws.merge_cells("A5:E5")
        ws["A5"] = "📞 ទូរស័ព្ទ: 096 29 647 13 | ✉️ អ៊ីមែល: dana267yue@gmail.com"
        ws["A5"].font = font_meta_lbl

        ws.merge_cells("A6:E6")
        ws["A6"] = f"📅 កាលបរិច្ឆេទបង្កើត: {now_str}"
        ws["A6"].font = font_meta_lbl

        ws.merge_cells("F4:G4")
        ws["F4"] = "អ្នកបង្កើត: Administrator"
        ws["F4"].font = font_meta_lbl
        ws["F4"].alignment = align_right

        total_vips = len(vip_qs)
        total_vip_orders = sum(v['order_count'] or 0 for v in vip_qs)
        total_vip_spend = sum(v['total_spent'] or 0 for v in vip_qs)

        kpi_cards = [
            ("A8:B8", "A9:B9", "A8", "A9", "ចំនួនអតិថិជន VIP", total_vips, "0F172A"),
            ("C8:D8", "C9:D9", "C8", "C9", "ការបញ្ជាទិញសរុប", total_vip_orders, "16A34A"),
            ("E8:G8", "E9:G9", "E8", "E9", "ចំណាយសរុបពី VIP ($)", f"${total_vip_spend:.2f}", "16A34A")
        ]

        ws.row_dimensions[8].height = 20
        ws.row_dimensions[9].height = 28

        for lbl_range, val_range, lbl_cell, val_cell, label_text, val_text, color_hex in kpi_cards:
            if ":" in lbl_range:
                ws.merge_cells(lbl_range)
                ws.merge_cells(val_range)
            ws[lbl_cell] = label_text
            ws[lbl_cell].font = font_kpi_lbl
            ws[lbl_cell].alignment = align_center
            ws[val_cell] = val_text
            ws[val_cell].font = Font(name="Khmer OS Siemreap", size=13, bold=True, color=color_hex)
            ws[val_cell].alignment = align_center

        for r in range(8, 10):
            for c in range(1, 8):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_kpi_card
                cell.border = cell_border

        headers = ["ល.រ", "ឈ្មោះអតិថិជន", "អ៊ីមែល", "លេខទូរស័ព្ទ", "ចំនួនបញ្ជាទិញ", "ចំណាយសរុប ($)", "បញ្ជាទិញចុងក្រោយ"]
        ws.row_dimensions[11].height = 34
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col_num, value=header)
            cell.font = font_tbl_header
            cell.fill = fill_tbl_header
            cell.alignment = align_center
            cell.border = cell_border

        current_row = 12
        for idx, item in enumerate(vip_qs, 1):
            cust_name = f"{item['last_name']} {item['first_name']}".strip() or item['user__username'] or "Guest"
            if len(cust_name) > 30 or len(str(item['email'] or "")) > 30:
                ws.row_dimensions[current_row].height = 32
            else:
                ws.row_dimensions[current_row].height = 26

            c1 = ws.cell(row=current_row, column=1, value=idx)
            c1.alignment = align_center

            c2 = ws.cell(row=current_row, column=2, value=cust_name)
            c2.alignment = align_left

            c3 = ws.cell(row=current_row, column=3, value=item['email'] or "-")
            c3.alignment = align_left

            c4 = ws.cell(row=current_row, column=4, value=item['phone'] or "-")
            c4.alignment = align_center

            c5 = ws.cell(row=current_row, column=5, value=item['order_count'])
            c5.alignment = align_center

            c6 = ws.cell(row=current_row, column=6, value=float(item['total_spent'] or 0))
            c6.number_format = "$#,##0.00"
            c6.alignment = align_right

            c7 = ws.cell(row=current_row, column=7, value=item['last_order'].strftime("%d/%m/%Y %H:%M") if item['last_order'] else "-")
            c7.alignment = align_center

            row_fill = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
            for c in range(1, 8):
                cell = ws.cell(row=current_row, column=c)
                if row_fill.fill_type: cell.fill = row_fill
                cell.border = cell_border
                if c != 6: cell.font = font_data
                else: cell.font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="16A34A")

            current_row += 1

        footer_row = current_row + 3
        ws.merge_cells(f"B{footer_row}:C{footer_row}")
        ws.cell(row=footer_row, column=2, value="អ្នករៀបចំ ៖ ____________________").font = font_footer
        ws.cell(row=footer_row, column=2).alignment = align_left

        ws.merge_cells(f"E{footer_row}:G{footer_row}")
        ws.cell(row=footer_row, column=5, value="អ្នកអនុម័ត ៖ ____________________").font = font_footer
        ws.cell(row=footer_row, column=5).alignment = align_right

        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 26
        ws.column_dimensions['C'].width = 28
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 20

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="PNK_VIP_Customers_Report_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response
    else:
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="PNK_VIP_Customers_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
        response.write('\ufeff')
        writer = csv.writer(response)
        writer.writerow(["ល.រ", "ឈ្មោះអតិថិជន", "អ៊ីមែល", "លេខទូរស័ព្ទ", "ចំនួនបញ្ជាទិញ", "ចំណាយសរុប ($)", "បញ្ជាទិញចុងក្រោយ"])
        for idx, item in enumerate(vip_qs, 1):
            cust_name = f"{item['last_name']} {item['first_name']}".strip() or item['user__username'] or "Guest"
            writer.writerow([idx, cust_name, item['email'] or "-", item['phone'] or "-", item['order_count'], f"{item['total_spent']:.2f}", item['last_order'].strftime("%d/%m/%Y %H:%M") if item['last_order'] else "-"])
        return response

@login_required
@user_passes_test(is_admin)
def export_low_stock_csv(request):
    low_stock_qs = Product.objects.filter(stock__lt=5).order_by('stock')

    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Low Stock Report"
        ws.views.sheetView[0].showGridLines = True

        # Page Setup (Landscape A4 + Fit All Columns on 1 Page Width)
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.6
        ws.page_margins.bottom = 0.6
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

        ws.oddHeader.left.text = "PNK SHOP - របាយការណ៍ស្តុកជិតអស់"
        ws.oddHeader.right.text = "កាលបរិច្ឆេទ: &[Date]"
        ws.oddFooter.left.text = "PNK SHOP"
        ws.oddFooter.center.text = "ទំព័រ &[Page] នៃ &[Pages]"

        font_title = Font(name="Khmer OS Muol Light", size=15, bold=True, color="FFFFFF")
        font_subtitle = Font(name="Khmer OS Battambang", size=11, bold=True, color="E2E8F0")
        font_meta_lbl = Font(name="Khmer OS Siemreap", size=10, bold=True, color="334155")
        font_meta_val = Font(name="Khmer OS Siemreap", size=10, color="0F172A")
        font_kpi_lbl = Font(name="Khmer OS Battambang", size=9, bold=True, color="475569")
        font_tbl_header = Font(name="Khmer OS Battambang", size=10.5, bold=True, color="FFFFFF")
        font_data = Font(name="Khmer OS Siemreap", size=10, color="0F172A")
        font_footer = Font(name="Khmer OS Siemreap", size=9.5, italic=True, color="475569")

        fill_header_banner = PatternFill(start_color="198754", end_color="146C43", fill_type="solid")
        fill_tbl_header = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_kpi_card = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        thin_border = Side(style='thin', color='CBD5E1')
        cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

        ws.merge_cells("A1:G1")
        ws["A1"] = "PNK SHOP"
        ws["A1"].font = font_title
        ws["A1"].alignment = align_center

        ws.merge_cells("A2:G2")
        ws["A2"] = "របាយការណ៍ទំនិញជិតអស់ពីស្តុក (Low Stock Alerts Report)"
        ws["A2"].font = font_subtitle
        ws["A2"].alignment = align_center

        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 26
        for r in range(1, 3):
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = fill_header_banner

        now_str = timezone.now().strftime('%d-%m-%Y %I:%M %p')
        ws.merge_cells("A4:E4")
        ws["A4"] = "📍 អាសយដ្ឋាន: ភូមិនិគមន៍លើ ខេត្តត្បូងឃ្មុំ, ព្រះរាជាណាចក្រកម្ពុជា"
        ws["A4"].font = font_meta_lbl

        ws.merge_cells("A5:E5")
        ws["A5"] = "📞 ទូរស័ព្ទ: 096 29 647 13 | ✉️ អ៊ីមែល: dana267yue@gmail.com"
        ws["A5"].font = font_meta_lbl

        ws.merge_cells("A6:E6")
        ws["A6"] = f"📅 កាលបរិច្ឆេទបង្កើត: {now_str}"
        ws["A6"].font = font_meta_lbl

        ws.merge_cells("F4:G4")
        ws["F4"] = "លក្ខខណ្ឌស្តុក: តិចជាង ៥ គ្រឿង"
        ws["F4"].font = font_meta_lbl
        ws["F4"].alignment = align_right

        count_low = low_stock_qs.filter(stock__gt=0).count()
        count_zero = low_stock_qs.filter(stock=0).count()
        stock_value = sum(p.stock * float(p.price) for p in low_stock_qs)

        kpi_cards = [
            ("A8:B8", "A9:B9", "A8", "A9", "ជិតអស់ស្តុក (< 5)", count_low, "D97706"),
            ("C8:D8", "C9:D9", "C8", "C9", "អស់ពីស្តុក (0)", count_zero, "DC2626"),
            ("E8:G8", "E9:G9", "E8", "E9", "តម្លៃស្តុកនៅសល់ ($)", f"${stock_value:.2f}", "0F172A")
        ]

        ws.row_dimensions[8].height = 20
        ws.row_dimensions[9].height = 28

        for lbl_range, val_range, lbl_cell, val_cell, label_text, val_text, color_hex in kpi_cards:
            if ":" in lbl_range:
                ws.merge_cells(lbl_range)
                ws.merge_cells(val_range)
            ws[lbl_cell] = label_text
            ws[lbl_cell].font = font_kpi_lbl
            ws[lbl_cell].alignment = align_center
            ws[val_cell] = val_text
            ws[val_cell].font = Font(name="Khmer OS Siemreap", size=13, bold=True, color=color_hex)
            ws[val_cell].alignment = align_center

        for r in range(8, 10):
            for c in range(1, 8):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_kpi_card
                cell.border = cell_border

        headers = ["ល.រ", "ID ទំនិញ", "ឈ្មោះទូរស័ព្ទ/ផលិតផល", "ម៉ាក (Brand)", "ស្តុកនៅសល់ (គ្រឿង)", "តម្លៃរាយ ($)", "ស្ថានភាពស្តុក"]
        ws.row_dimensions[11].height = 34
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col_num, value=header)
            cell.font = font_tbl_header
            cell.fill = fill_tbl_header
            cell.alignment = align_center
            cell.border = cell_border

        current_row = 12
        for idx, item in enumerate(low_stock_qs, 1):
            pname = str(item.name or "")
            if len(pname) > 40:
                ws.row_dimensions[current_row].height = 42
            elif len(pname) > 22:
                ws.row_dimensions[current_row].height = 32
            else:
                ws.row_dimensions[current_row].height = 26

            c1 = ws.cell(row=current_row, column=1, value=idx)
            c1.alignment = align_center

            c2 = ws.cell(row=current_row, column=2, value=f"#{item.id}")
            c2.alignment = align_center

            c3 = ws.cell(row=current_row, column=3, value=item.name)
            c3.alignment = align_left

            c4 = ws.cell(row=current_row, column=4, value=item.brand.name if item.brand else "-")
            c4.alignment = align_center

            c5 = ws.cell(row=current_row, column=5, value=item.stock)
            c5.alignment = align_center

            c6 = ws.cell(row=current_row, column=6, value=float(item.price))
            c6.number_format = "$#,##0.00"
            c6.alignment = align_right

            status_txt = "អស់ពីស្តុក" if item.stock == 0 else f"សល់ {item.stock} គ្រឿង"
            c7 = ws.cell(row=current_row, column=7, value=status_txt)
            c7.alignment = align_center

            row_fill = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
            for c in range(1, 8):
                cell = ws.cell(row=current_row, column=c)
                if row_fill.fill_type: cell.fill = row_fill
                cell.border = cell_border
                cell.font = font_data

            if item.stock == 0:
                ws.cell(row=current_row, column=7).font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="DC2626")
            else:
                ws.cell(row=current_row, column=7).font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="D97706")

            current_row += 1

        footer_row = current_row + 3
        ws.merge_cells(f"B{footer_row}:C{footer_row}")
        ws.cell(row=footer_row, column=2, value="អ្នករៀបចំ ៖ ____________________").font = font_footer
        ws.cell(row=footer_row, column=2).alignment = align_left

        ws.merge_cells(f"E{footer_row}:G{footer_row}")
        ws.cell(row=footer_row, column=5, value="អ្នកអនុម័ត ៖ ____________________").font = font_footer
        ws.cell(row=footer_row, column=5).alignment = align_right

        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 38
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 18

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="PNK_Low_Stock_Report_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response
    else:
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="PNK_Low_Stock_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
        response.write('\ufeff')
        writer = csv.writer(response)
        writer.writerow(["ល.រ", "ID ទំនិញ", "ឈ្មោះទូរស័ព្ទ/ផលិតផល", "ម៉ាក (Brand)", "ស្តុកនៅសល់ (គ្រឿង)", "តម្លៃរាយ ($)", "ស្ថានភាពស្តុក"])
        for idx, item in enumerate(low_stock_qs, 1):
            status_txt = "អស់ពីស្តុក" if item.stock == 0 else f"សល់ {item.stock} គ្រឿង"
            writer.writerow([idx, f"#{item.id}", item.name, item.brand.name if item.brand else "-", item.stock, f"{item.price:.2f}", status_txt])
        return response


@login_required
@user_passes_test(is_admin)
def customer_messages(request):
    customer_qs = User.objects.filter(is_staff=False, is_superuser=False)
    
    # If no messages exist in system at all, seed realistic initial customer messages
    if not ChatMessage.objects.exists() and customer_qs.exists():
        sample_questions = [
            "ជម្រាបសួរខណៈបង! ខ្ញុំបានឃើញហាងបងបង្ហោះលក់ទូរស័ព្ទ iPhone 15 Pro Max។ ចង់សួរថាតើឥឡូវនេះនៅសល់ពណ៌អ្វីខ្លះដែរបង?",
            "បង! តើហាងមានសេវាបង់រំលស់អត់បង? ហើយត្រូវការឯកសារអ្វីខ្លះ?",
            "ជម្រាបសួរ! តើហាងបើកលក់ពីម៉ោងប៉ុន្មានដល់ម៉ោងប៉ុន្មានដែរបង?",
            "បង មានម៉ាស៊ីនសំណុំហ្ស៊ីននៅសល់អត់? បើទិញដាច់មានថែមជូនកាដូអ្វីខ្លះ?",
            "អរគុណច្រើនបង! ខ្ញុំបានទទួលទូរស័ព្ទហើយ ស្អាតណាស់!"
        ]
        for idx, u in enumerate(customer_qs[:5]):
            q = sample_questions[idx % len(sample_questions)]
            ChatMessage.objects.create(
                user=u,
                sender='customer',
                message=q,
                is_read=False
            )
    
    customers = list(customer_qs)
    def get_sort_key(u):
        last_m = ChatMessage.objects.filter(user=u).order_by('-created_at').first()
        return last_m.created_at if last_m else u.date_joined

    customers.sort(key=get_sort_key, reverse=True)

    active_user_id = request.GET.get('user_id')
    active_user = None
    if active_user_id:
        active_user = customer_qs.filter(id=active_user_id).first()
    
    if not active_user and len(customers) > 0:
        latest_msg = ChatMessage.objects.order_by('-created_at').first()
        if latest_msg:
            active_user = latest_msg.user
        else:
            active_user = customers[0]

    customer_list = []
    for c in customers:
        last_msg = ChatMessage.objects.filter(user=c).order_by('-created_at').first()
        unread_cnt = ChatMessage.objects.filter(user=c, sender='customer', is_read=False).count()
        order_cnt = Order.objects.filter(user=c).count()
        phone = getattr(getattr(c, 'profile', None), 'phone', None) or Order.objects.filter(user=c).values_list('phone', flat=True).first() or "-"
        
        customer_list.append({
            'user': c,
            'last_msg': last_msg,
            'unread_cnt': unread_cnt,
            'order_cnt': order_cnt,
            'phone': phone,
        })
        
    messages_list = []
    if active_user:
        ChatMessage.objects.filter(user=active_user, sender='customer', is_read=False).update(is_read=True)
        messages_list = ChatMessage.objects.filter(user=active_user).order_by('created_at')

    active_user_order_cnt = Order.objects.filter(user=active_user).count() if active_user else 0
    active_user_phone = getattr(getattr(active_user, 'profile', None), 'phone', None) or (Order.objects.filter(user=active_user).values_list('phone', flat=True).first() if active_user else "-")

    context = {
        'customers': customer_list,
        'active_user': active_user,
        'messages_list': messages_list,
        'active_user_order_cnt': active_user_order_cnt,
        'active_user_phone': active_user_phone,
    }
    return render(request, 'accounts/dashboard/contact.html', context)


from django.views.decorators.http import require_POST
from django.http import JsonResponse

@login_required
@user_passes_test(is_admin)
@require_POST
def send_chat_message_api(request):
    user_id = request.POST.get('user_id')
    message_text = request.POST.get('message', '').strip()
    sender = request.POST.get('sender', 'admin')
    
    if not user_id or not message_text:
        return JsonResponse({'status': 'error', 'error': 'Missing user or message text.'}, status=400)
        
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'error': 'User not found.'}, status=404)
        
    msg = ChatMessage.objects.create(
        user=user,
        sender=sender,
        message=message_text,
        is_read=True if sender == 'admin' else False
    )
    
    return JsonResponse({
        'status': 'success',
        'id': msg.id,
        'user_id': user.id,
        'sender': msg.sender,
        'message': msg.message,
        'created_at': msg.created_at.strftime('%H:%M')
    })


@login_required
@user_passes_test(is_admin)
@require_POST
def update_auto_reply_settings_api(request):
    store_settings = StoreSetting.get_settings()
    
    chat_auto_reply = request.POST.get('chat_auto_reply', '').strip()
    enable_auto_reply = request.POST.get('enable_auto_reply') in ['true', 'True', '1', 'on']
    telegram_link = request.POST.get('telegram_link', '').strip()
    
    if chat_auto_reply:
        store_settings.chat_auto_reply = chat_auto_reply
    store_settings.enable_auto_reply = enable_auto_reply
    if telegram_link:
        store_settings.telegram_link = telegram_link
        
    store_settings.save()
    
    log_activity(
        request.user, 
        "បច្ចុប្បន្នភាពសារឆ្លើយតបស្វ័យប្រវត្តិ", 
        f"Auto-reply: {'បើក' if enable_auto_reply else 'បិទ'}", 
        "bi-robot", 
        "text-success"
    )
    
    return JsonResponse({
        'status': 'success',
        'message': 'បានរក្សាទុកការកំណត់សារឆ្លើយតបស្វ័យប្រវត្តិដោយជោគជ័យ!',
        'chat_auto_reply': store_settings.chat_auto_reply,
        'enable_auto_reply': store_settings.enable_auto_reply,
        'telegram_link': store_settings.telegram_link
    })


@require_POST
def customer_send_chat_api(request):
    message_text = request.POST.get('message', '').strip()
    if not message_text:
        return JsonResponse({'status': 'error', 'error': 'Missing message text.'}, status=400)

    if request.user.is_authenticated:
        user = request.user
    else:
        user, _ = User.objects.get_or_create(
            username='GuestCustomer',
            defaults={
                'first_name': 'អតិថិជន',
                'last_name': 'ភ្ញៀវ',
                'email': 'guest@pnkshop.com',
                'is_staff': False,
                'is_superuser': False
            }
        )

    msg = ChatMessage.objects.create(
        user=user,
        sender='customer',
        message=message_text,
        is_read=False
    )

    return JsonResponse({
        'status': 'success',
        'id': msg.id,
        'user_id': user.id,
        'username': user.username,
        'sender': msg.sender,
        'message': msg.message,
        'created_at': msg.created_at.strftime('%H:%M')
    })


def check_admin_online():
    try:
        from django.contrib.sessions.models import Session
        sessions = Session.objects.filter(expire_date__gte=timezone.now())
        user_ids = set()
        for s in sessions:
            data = s.get_decoded()
            uid = data.get('_auth_user_id')
            if uid:
                user_ids.add(uid)
        if user_ids and User.objects.filter(id__in=user_ids, is_staff=True).exists():
            return True
    except Exception:
        pass

    recent_cutoff = timezone.now() - timedelta(minutes=30)
    return User.objects.filter(is_staff=True, last_login__gte=recent_cutoff).exists()


def get_customer_chat_history_api(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        user = User.objects.filter(username='GuestCustomer').first()
        
    messages_data = []
    if user:
        msgs = ChatMessage.objects.filter(user=user).order_by('created_at')
        for m in msgs:
            messages_data.append({
                'id': m.id,
                'sender': m.sender,
                'message': m.message,
                'created_at': m.created_at.strftime('%H:%M')
            })
            
    return JsonResponse({
        'status': 'success',
        'messages': messages_data,
        'is_admin_online': check_admin_online()
    })


@login_required
@user_passes_test(is_admin)
def get_latest_messages_api(request):
    active_user_id = request.GET.get('user_id')
    last_msg_id = request.GET.get('last_msg_id', 0)
    
    new_messages = []
    if active_user_id:
        msgs = ChatMessage.objects.filter(user_id=active_user_id, id__gt=last_msg_id).order_by('created_at')
        for m in msgs:
            if m.sender == 'customer' and not m.is_read:
                m.is_read = True
                m.save()
            new_messages.append({
                'id': m.id,
                'user_id': m.user.id,
                'sender': m.sender,
                'message': m.message,
                'created_at': m.created_at.strftime('%H:%M')
            })

    latest_overall_msg = ChatMessage.objects.filter(sender='customer', id__gt=last_msg_id).order_by('-id').first()
    latest_info = None
    if latest_overall_msg:
        latest_info = {
            'id': latest_overall_msg.id,
            'user_id': latest_overall_msg.user.id,
            'username': latest_overall_msg.user.username,
            'message': latest_overall_msg.message,
            'created_at': latest_overall_msg.created_at.strftime('%H:%M')
        }

    total_unread = ChatMessage.objects.filter(sender='customer', is_read=False).count()
    return JsonResponse({
        'status': 'success',
        'new_messages': new_messages,
        'latest_info': latest_info,
        'total_unread': total_unread
    })


def admin_heartbeat_api(request):
    """
    Heartbeat API សម្រាប់ Update Session Activity ពេល Admin/Staff កំពុងធ្វើការងារ
    """
    if request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser):
        import time
        request.session['admin_last_activity'] = int(time.time())
        return JsonResponse({'status': 'active', 'authenticated': True})
    return JsonResponse({'status': 'unauthorized', 'authenticated': False}, status=401)


def get_notifications_api(request):
    """
    Real-time Live Notification API សម្រាប់ Refresh សារជូនដំណឹងលើកណ្ដឹងដោយមិនចាំបាច់ Reload ទំព័រ
    """
    if not (request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser)):
        return JsonResponse({'status': 'unauthorized'}, status=401)
    
    from accounts.context_processors import dashboard_notifications
    from django.template.loader import render_to_string
    
    notif_context = dashboard_notifications(request)
    html_content = render_to_string('accounts/dashboard/includes/notification_items.html', notif_context, request=request)
    
    return JsonResponse({
        'status': 'success',
        'total_count': notif_context['total_notifications_count'],
        'pending_orders_count': notif_context['pending_orders_count'],
        'low_stock_count': notif_context['low_stock_count'],
        'new_customers_count': notif_context['new_customers_count'],
        'unread_chat_count': notif_context['unread_chat_count'],
        'html': html_content
    })


def mark_notifications_read_api(request):
    """
    API សម្រាប់សម្គាល់ថាបានមើលការជូនដំណឹង (Mark notifications as read)
    """
    if not (request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser)):
        return JsonResponse({'status': 'unauthorized'}, status=401)
    
    from django.contrib.auth.models import User
    notif_type = request.GET.get('type', 'customers')
    
    if notif_type in ['customers', 'all']:
        latest_cust = User.objects.filter(is_staff=False, is_superuser=False).order_by('-id').first()
        if latest_cust:
            request.session['last_seen_customer_id'] = latest_cust.id
            request.session.modified = True
            
    from accounts.context_processors import dashboard_notifications
    notif_context = dashboard_notifications(request)
    
    return JsonResponse({
        'status': 'success',
        'total_count': notif_context['total_notifications_count'],
        'new_customers_count': notif_context['new_customers_count']
    })